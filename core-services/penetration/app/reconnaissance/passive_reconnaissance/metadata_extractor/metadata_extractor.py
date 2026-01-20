import os
import PyPDF2
from datetime import datetime, timezone # For parsing PDF dates
from openpyxl import load_workbook # For XLSX
from PIL import Image, UnidentifiedImageError # For images, import specific error
from PIL.ExifTags import TAGS # For images
from logging_config import get_logger
logger = get_logger(__name__)

# Attempt to import docx, handle if not present
try:
    from docx import Document as DocxDocument
    HAS_PYTHON_DOCX = True
except ImportError:
    HAS_PYTHON_DOCX = False
    # This print might be noisy if the module is imported elsewhere.
    # Consider logging or a status flag instead for library use.
    # print("Warning: python-docx library not found. DOCX metadata extraction will be disabled.")

def parse_pdf_date_string(date_str):
    if not date_str or not isinstance(date_str, str):
        return None
    
    original_date_str = date_str
    if date_str.startswith("D:"):
        date_str = date_str[2:]

    # Remove timezone offset like Z, +HH'MM', -HH'MM' for strptime
    # and store it if possible. This is a simplified handling.
    tzinfo_obj = None
    if date_str.endswith("Z"):
        date_str = date_str[:-1]
        tzinfo_obj = timezone.utc
    elif '+' in date_str[14:]: # Check for '+' in timezone part
        try:
            main_part, offset_part = date_str.rsplit('+', 1)
            if "'" in offset_part:
                offset_part = offset_part.replace("'", "")
                if len(offset_part) == 4: # HHMM
                    h, m = int(offset_part[:2]), int(offset_part[2:])
                    tzinfo_obj = timezone(datetime.timedelta(hours=h, minutes=m))
                    date_str = main_part
        except ValueError: pass # Could not parse offset
    elif '-' in date_str[14:]: # Check for '-' in timezone part
        try:
            main_part, offset_part = date_str.rsplit('-', 1)
            if "'" in offset_part:
                offset_part = offset_part.replace("'", "")
                if len(offset_part) == 4: # HHMM
                    h, m = int(offset_part[:2]), int(offset_part[2:])
                    tzinfo_obj = timezone(datetime.timedelta(hours=-h, minutes=-m))
                    date_str = main_part
        except ValueError: pass

    # Try parsing the common YYYYMMDDHHMMSS format
    # Some PDF dates might only have YYYYMMDD or YYYYMMDDHHMM
    possible_formats = ["%Y%m%d%H%M%S", "%Y%m%d%H%M", "%Y%m%d"]
    parsed_dt = None
    for fmt in possible_formats:
        try:
            parsed_dt = datetime.strptime(date_str[:len(fmt)], fmt) # Parse only the expected length
            if tzinfo_obj:
                parsed_dt = parsed_dt.replace(tzinfo=tzinfo_obj)
            return parsed_dt
        except ValueError:
            continue
    return original_date_str # Fallback to original string if all parsing fails

def extract_pdf_metadata(file_path):
    metadata = {}
    try:
        with open(file_path, 'rb') as pdf_file:
            reader = PyPDF2.PdfReader(pdf_file)
            doc_info = reader.metadata
            if doc_info:
                metadata = {
                    "title": doc_info.title, "author": doc_info.author,
                    "subject": doc_info.subject, "creator": doc_info.creator,
                    "producer": doc_info.producer,
                    "creation_date": parse_pdf_date_string(doc_info.creation_date),
                    "modification_date": parse_pdf_date_string(doc_info.modification_date),
                    "page_count": len(reader.pages), "is_encrypted": reader.is_encrypted,
                }
            else:
                metadata["message"] = "No metadata found in PDF."
    except PyPDF2.errors.PdfReadError as e:
        return {"error": f"PyPDF2 could not read PDF (corrupted/encrypted?): {e}"}
    except Exception as e:
        return {"error": f"Error extracting PDF metadata from '{os.path.basename(file_path)}': {e}"}
    return metadata

def extract_docx_metadata(file_path):
    if not HAS_PYTHON_DOCX:
        return {"error": "python-docx library not installed. DOCX extraction skipped."}
    metadata = {}
    try:
        doc = DocxDocument(file_path)
        cp = doc.core_properties
        metadata = {
            "title": cp.title, "author": cp.author, "subject": cp.subject,
            "keywords": cp.keywords, "comments": cp.comments,
            "last_modified_by": cp.last_modified_by, "revision": cp.revision,
            "created_date": cp.created.isoformat() if cp.created else None,
            "modified_date": cp.modified.isoformat() if cp.modified else None,
            "category": cp.category, "language": cp.language,
        }
    except Exception as e:
        return {"error": f"Error extracting DOCX metadata from '{os.path.basename(file_path)}': {e}"}
    return metadata

def extract_xlsx_metadata(file_path):
    metadata = {}
    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        cp = workbook.properties
        metadata = {
            "title": cp.title, "creator": cp.creator, "subject": cp.subject,
            "description": cp.description, "keywords": cp.keywords,
            "last_modified_by": cp.lastModifiedBy,
            "created_date": cp.created.isoformat() if cp.created and isinstance(cp.created, datetime) else cp.created,
            "modified_date": cp.modified.isoformat() if cp.modified and isinstance(cp.modified, datetime) else cp.modified,
            "category": cp.category, "language": cp.language,
            "sheet_names": workbook.sheetnames or None,
        }
    except Exception as e:
        return {"error": f"Error extracting XLSX metadata from '{os.path.basename(file_path)}': {e}"}
    return metadata

def extract_image_metadata(file_path):
    metadata = {"exif": None} # Initialize EXIF as None
    try:
        image = Image.open(file_path)
        metadata["format"] = image.format
        metadata["mode"] = image.mode
        metadata["size"] = image.size

        exif_data_raw = image._getexif() # Returns dict or None
        if exif_data_raw:
            metadata["exif"] = {} # Create dict if EXIF data exists
            for tag_id, value in exif_data_raw.items():
                tag_name = TAGS.get(tag_id, str(tag_id)) # Use string of tag_id if name not found
                if isinstance(value, bytes):
                    try: value = value.decode('utf-8', errors='replace')
                    except: value = str(value)
                metadata["exif"][str(tag_name)] = value
        # Consider image.info for other metadata if needed
        # metadata["pil_info"] = dict(image.info)
    except FileNotFoundError:
        return {"error": f"Image file not found: {file_path}"}
    except UnidentifiedImageError:
         return {"error": f"Cannot identify image file (corrupted/unsupported): {os.path.basename(file_path)}."}
    except Exception as e:
        return {"error": f"Error extracting image metadata from '{os.path.basename(file_path)}': {e}"}
    return metadata

def extract_metadata(file_path):
    if not os.path.exists(file_path):
        return {"error": "File does not exist", "file_path": file_path}
    if not os.path.isfile(file_path):
        return {"error": "Path is not a file", "file_path": file_path}

    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    # Basic file system metadata common to all files
    try:
        common_meta = {
            "file_name": os.path.basename(file_path), "file_extension": ext,
            "file_size_bytes": os.path.getsize(file_path),
            "last_accessed_time_utc": datetime.fromtimestamp(os.path.getatime(file_path), timezone.utc).isoformat(),
            "last_modified_time_utc": datetime.fromtimestamp(os.path.getmtime(file_path), timezone.utc).isoformat(),
            "creation_time_utc": datetime.fromtimestamp(os.path.getctime(file_path), timezone.utc).isoformat(),
        }
    except Exception as e: # Handle potential OS errors for stat calls
        return {"error": f"Could not retrieve basic file system metadata for {file_path}: {e}"}

    type_specific_meta = None
    if ext == ".pdf": type_specific_meta = extract_pdf_metadata(file_path)
    elif ext == ".docx": type_specific_meta = extract_docx_metadata(file_path)
    elif ext == ".xlsx": type_specific_meta = extract_xlsx_metadata(file_path)
    elif ext in [".jpg", ".jpeg", ".png", ".tiff", ".gif", ".bmp"]:
        type_specific_meta = extract_image_metadata(file_path)
    else:
        common_meta["warning"] = f"Unsupported file type for detailed metadata: {ext}"
        return common_meta

    # Merge, prioritizing error from specific extractor
    if type_specific_meta and "error" in type_specific_meta:
        common_meta.update(type_specific_meta) # This adds the error key
        return common_meta

    if type_specific_meta:
        common_meta.update(type_specific_meta)

    return common_meta

if __name__ == "__main__":
    # For local testing, create dummy files if they don't exist.
    # This setup code should ideally not be in the final library.
    dummy_files_created = {}
    def setup_dummy_file(name, creation_func):
        if not os.path.exists(name):
            try:
                creation_func()
                dummy_files_created[name] = True
                logger.info(f"Created dummy file: {name}")
            except Exception as e:
                logger.info(f"Could not create dummy file {name}: {e}. Test for it will be skipped.")
                dummy_files_created[name] = False
        else:
            dummy_files_created[name] = True # Assume it's usable if it exists

    # Dummy file creation functions
    def create_dummy_pdf():
        from reportlab.pdfgen import canvas
        c = canvas.Canvas("dummy.pdf"); c.setTitle("Dummy PDF"); c.setAuthor("Test Author"); c.save()
    def create_dummy_docx():
        if HAS_PYTHON_DOCX:
            doc = DocxDocument(); doc.core_properties.title="Dummy DOCX"; doc.save("dummy.docx")
    def create_dummy_xlsx():
        from openpyxl import Workbook
        wb = Workbook(); wb.properties.title="Dummy XLSX"; wb.save("dummy.xlsx")
    def create_dummy_jpg():
        Image.new('RGB', (10, 10), color='blue').save("dummy.jpg")

    setup_dummy_file("dummy.pdf", create_dummy_pdf)
    setup_dummy_file("dummy.docx", create_dummy_docx)
    setup_dummy_file("dummy.xlsx", create_dummy_xlsx)
    setup_dummy_file("dummy.jpg", create_dummy_jpg)

    test_file_paths = ["dummy.pdf", "dummy.docx", "dummy.xlsx", "dummy.jpg", "nonexistent.file"]

    for path in test_file_paths:
        logger.info(f"\n--- Metadata for: {path} ---")
        if "dummy" in path and not dummy_files_created.get(path, False):
            logger.info(f"  Skipping test for {path} as it was not created/found.")
            continue

        result = extract_metadata(path)
        for key, value in result.items():
            if key == "exif" and isinstance(value, dict):
                logger.info(f"  Exif Data:")
                for exif_k, exif_v in value.items():
                    logger.info(f"    - {exif_k}: {str(exif_v)[:100]}") # Truncate long EXIF values
            elif isinstance(value, datetime):
                logger.info(f"  {key.replace('_',' ').title()}: {value.isoformat()}")
            else:
                logger.info(f"  {key.replace('_',' ').title()}: {value}")
        logger.info("-" * 40)

    # Optional: Clean up dummy files after testing
    # for f_name in ["dummy.pdf", "dummy.docx", "dummy.xlsx", "dummy.jpg"]:
    #     if os.path.exists(f_name) and dummy_files_created.get(f_name): os.remove(f_name)
    # print("\nCleaned up dummy files if they were created by this script.")
